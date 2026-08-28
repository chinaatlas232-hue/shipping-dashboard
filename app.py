import datetime
import io
import os
import openpyxl
import pandas as pd
import streamlit as st

st.set_page_config(page_title="وصل تسليم بضاعة - أطلس", layout="wide")

# --- مسارات الملفات ---
UPLOAD_DIR = "saved_files"
os.makedirs(UPLOAD_DIR, exist_ok=True)

shipment_path = os.path.join(UPLOAD_DIR, "shipments_data.xlsx")
template_path = os.path.join(UPLOAD_DIR, "template.xlsx")
logo_path = os.path.join(UPLOAD_DIR, "logo.png")
customer_info_path = os.path.join(UPLOAD_DIR, "customer_info.xlsx")

# --- تنسيق الشريط الجانبي ---
st.markdown(
    """
    <style>
    [data-testid="stSidebar"] {
        background-color: #334155;
        color: #f8fafc;
    }
    [data-testid="stSidebar"] h1, 
    [data-testid="stSidebar"] h2, 
    [data-testid="stSidebar"] h3, 
    [data-testid="stSidebar"] label, 
    [data-testid="stSidebar"] .stMarkdown {
        color: #f8fafc !important;
    }
    [data-testid="stSidebar"] button[kind="secondary"] {
        background-color: #991b1b !important;
        color: white !important;
        border: 1px solid #7f1d1d !important;
        font-weight: bold !important;
    }
    </style>
""",
    unsafe_allow_html=True,
)

st.title("📦 النظام المالي والفني - وصل تسليم البضائع (شركة أطلس المحيط)")

with st.sidebar:
  st.header("⚙️ إدارة الملفات")

  uploaded_data_file = st.file_uploader(
      "1. ملف بيانات الشحنات (تعبئة وصل اطلس.xlsx)", type=["xlsx"]
  )
  if uploaded_data_file is not None:
    with open(shipment_path, "wb") as f:
      f.write(uploaded_data_file.getbuffer())
    st.success("تم حفظ ملف الشحنات بنجاح!")

  uploaded_template_file = st.file_uploader(
      "2. قالب وصل التسليم (Atlas_Cargo_Delivery_Receipt.xlsx)", type=["xlsx"]
  )
  if uploaded_template_file is not None:
    with open(template_path, "wb") as f:
      f.write(uploaded_template_file.getbuffer())
    st.success("تم حفظ القالب بنجاح!")

  uploaded_logo = st.file_uploader(
      "3. شعار الشركة (Logo)", type=["png", "jpg", "jpeg"]
  )
  if uploaded_logo is not None:
    with open(logo_path, "wb") as f:
      f.write(uploaded_logo.getbuffer())
    st.success("تم حفظ الشعار بنجاح!")

  uploaded_customer_file = st.file_uploader(
      "4. ملف معلومات العملاء (coustmer info)", type=["xlsx", "csv"]
  )
  if uploaded_customer_file is not None:
    with open(customer_info_path, "wb") as f:
      f.write(uploaded_customer_file.getbuffer())
    st.success("تم حفظ ملف معلومات العملاء بنجاح!")

  if st.button("🗑️ مسح الذاكرة ورفع ملفات جديدة"):
    for path in [shipment_path, template_path, logo_path, customer_info_path]:
      if os.path.exists(path):
        os.remove(path)
    st.rerun()

  # --- دالة الدمج وتحديث الأعمدة الإجبارية ---
  def load_and_merge_data():
    if not os.path.exists(shipment_path):
      return None

    df_s = pd.read_excel(shipment_path)
    df_s.columns = df_s.columns.astype(str).str.strip()

    if os.path.exists(customer_info_path):
      try:
        if customer_info_path.endswith('.csv'):
          df_c = pd.read_csv(customer_info_path)
        else:
          df_c = pd.read_excel(customer_info_path)
          
        df_c.columns = df_c.columns.astype(str).str.strip()

        ship_code_col = next((c for c in df_s.columns if "كود" in c or "code" in c.lower()), "الكود")
        cust_code_col = next((c for c in df_c.columns if "كود" in c or "code" in c.lower()), "الكود")

        if ship_code_col in df_s.columns and cust_code_col in df_c.columns:
          df_s['__s_code__'] = df_s[ship_code_col].astype(str).str.strip().str.upper().str.replace('.0', '', regex=False)
          df_c['__c_code__'] = df_c[cust_code_col].astype(str).str.strip().str.upper().str.replace('.0', '', regex=False)
          
          c_name_col = next((c for c in df_c.columns if 'الاسم' in c or 'name' in c.lower()), None)
          c_phone_col = next((c for c in df_c.columns if 'هاتف' in c or 'عاتف' in c or 'phone' in c.lower()), None)
          c_addr_col = next((c for c in df_c.columns if 'عنوان' in c or 'address' in c.lower()), None)

          name_dict = dict(zip(df_c['__c_code__'], df_c[c_name_col])) if c_name_col else {}
          phone_dict = dict(zip(df_c['__c_code__'], df_c[c_phone_col])) if c_phone_col else {}
          addr_dict = dict(zip(df_c['__c_code__'], df_c[c_addr_col])) if c_addr_col else {}

          s_name_col = next((c for c in df_s.columns if 'الاسم' in c and c != '__s_code__'), 'الاسم')
          s_phone_col = next((c for c in df_s.columns if 'هاتف' in c or 'عاتف' in c), 'رقم الهاتف')
          s_addr_col = next((c for c in df_s.columns if 'عنوان' in c), 'عنوان استلام البظاعة')

          df_s[s_name_col] = df_s['__s_code__'].map(name_dict).fillna(df_s.get(s_name_col))
          df_s[s_phone_col] = df_s['__s_code__'].map(phone_dict).fillna(df_s.get(s_phone_col))
          df_s[s_addr_col] = df_s['__s_code__'].map(addr_dict).fillna(df_s.get(s_addr_col))

          df_s.drop(columns=['__s_code__'], errors='ignore', inplace=True)
      except Exception as e:
        print(f"Merge error: {e}")
        
    return df_s

  # --- الفلاتر ---
  st.markdown("---")
  st.header("🔍 فلتر الشحنات")
  selected_shipment_filter = "الكل"
  selected_code_filter = "الكل"
  selected_type_filter = "الكل"

  temp_df = load_and_merge_data()
  if temp_df is not None and not temp_df.empty:
    try:
      ship_col = next((c for c in temp_df.columns if "شحنة" in str(c) or "shipment" in str(c).lower()), temp_df.columns[0])
      temp_df[ship_col] = temp_df[ship_col].fillna("بدون شحنة").astype(str).str.replace(".0", "", regex=False)
      shipment_list = ["الكل"] + sorted(temp_df[ship_col].unique().tolist())
      selected_shipment_filter = st.selectbox("اختر الشحنة للعرض:", shipment_list)
        
      filtered_temp_df = temp_df.copy()
      if selected_shipment_filter != "الكل":
        filtered_temp_df = filtered_temp_df[filtered_temp_df[ship_col] == selected_shipment_filter]

      code_col = next((c for c in filtered_temp_df.columns if "كود" in str(c) or "code" in str(c).lower()), None)
      if code_col:
        filtered_temp_df[code_col] = filtered_temp_df[code_col].fillna("بدون كود").astype(str).str.replace(".0", "", regex=False)
        code_list = ["الكل"] + sorted(filtered_temp_df[code_col].unique().tolist())
        selected_code_filter = st.selectbox("اختر أو ابحث برقم الكود:", code_list)
        if selected_code_filter != "الكل":
          filtered_temp_df = filtered_temp_df[filtered_temp_df[code_col] == selected_code_filter]

      type_col = next((c for c in filtered_temp_df.columns if "نوع" in str(c) or "type" in str(c).lower()), None)
      if type_col:
        filtered_temp_df[type_col] = filtered_temp_df[type_col].fillna("غير محدد").astype(str).str.strip()
        type_list = ["الكل"] + sorted(filtered_temp_df[type_col].unique().tolist())
        selected_type_filter = st.selectbox("اختر نوع الشحنة:", type_list)
    except Exception:
      pass

active_data_file = shipment_path if os.path.exists(shipment_path) else None
active_template_file = template_path if os.path.exists(template_path) else None
active_logo = logo_path if os.path.exists(logo_path) else None

if active_data_file is not None and active_template_file is not None:
  try:
    df = load_and_merge_data()
    if df is None or df.empty:
      st.warning("⚠️ ملف البيانات فارغ.")
      st.stop()

    ship_col = next((c for c in df.columns if "شحنة" in str(c) or "shipment" in str(c).lower()), df.columns[0])
    code_col = next((c for c in df.columns if "كود" in str(c) or "code" in str(c).lower()), None)
    type_col_name = next((c for c in df.columns if "نوع" in str(c) or "type" in str(c).lower()), None)

    df[ship_col] = df[ship_col].fillna("بدون شحنة").astype(str).str.replace(".0", "", regex=False)
    if code_col:
      df[code_col] = df[code_col].fillna("بدون كود").astype(str).str.replace(".0", "", regex=False)
    if type_col_name:
      df[type_col_name] = df[type_col_name].fillna("غير محدد").astype(str).str.strip()

    if selected_shipment_filter != "الكل":
      df = df[df[ship_col] == selected_shipment_filter]

    if selected_code_filter != "الكل" and code_col:
      df = df[df[code_col] == selected_code_filter]

    if selected_type_filter != "الكل" and type_col_name:
      df = df[df[type_col_name] == selected_type_filter]

    if df.empty:
      st.warning("⚠️ لا توجد بيانات مطابقة للفلاتر المحددة.")
      st.stop()

    today_date = datetime.date.today().strftime("%Y-%m-%d")

    import base64
    logo_base64 = ""
    if active_logo and os.path.exists(active_logo):
      with open(active_logo, "rb") as img_file:
        logo_base64 = base64.b64encode(img_file.read()).decode("utf-8")

    st.success(f"✅ تمت المطابقة بنجاح. الشحنة: **{selected_shipment_filter}** | الكود: **{selected_code_filter}** | النوع: **{selected_type_filter}**")
    st.markdown("---")

    total_clients_count = len(df)
    total_packages_count = 0
    total_weight_sum = 0.0
    total_cbm_sum = 0.0
    total_sales_sum = 0.0

    receipts_data_list = []
    all_receipts_html_for_print = ""

    for index, row in df.iterrows():
      shipment = str(row.get(ship_col, "بدون شحنة")).strip()
      code = str(row.get(code_col, "بدون كود")).strip() if code_col else ""
      
      display_code = "" if code in ["بدون كود", "nan", "None"] else code
      display_shipment = "" if shipment in ["بدون شحنة", "nan", "None"] else shipment

      name_col = next((c for c in df.columns if 'الاسم' in c or 'name' in c.lower()), None)
      name = str(row.get(name_col, "عميل غير محدد")).strip() if name_col else "عميل غير محدد"
      if name in ["nan", "None", ""]:
        name = "عميل غير محدد"

      file_name_id = f"Shipment_{shipment}_Client_{name}".replace(" ", "_")

      weight_col = next((c for c in df.columns if "وزن" in c or "weight" in c.lower()), None)
      weight = float(row.get(weight_col, 0) or 0) if weight_col else 0.0
      total_weight_sum += weight

      cbm_value = 0.0
      cbm_col = next((c for c in df.columns if "cbm" in c.lower() or "حجم" in c), None)
      if cbm_col:
        try:
          cbm_value = float(row.get(cbm_col, 0) or 0)
        except:
          pass
      total_cbm_sum += cbm_value

      packages_col = next((c for c in df.columns if "طرود" in c or "packages" in c.lower()), None)
      try:
        packages = int(float(row.get(packages_col, 0) or 0)) if packages_col else 0
      except:
        packages = 0
      total_packages_count += packages

      price_col = next((c for c in df.columns if "سعر" in c or "price" in c.lower()), None)
      price_per_kg = float(row.get(price_col, 0) or 0) if price_col else 0.0

      sales_col = next((c for c in df.columns if "مبيعات" in c or "اجمالي" in c or "total" in c.lower()), None)
      total_sales = float(row.get(sales_col, 0) or 0) if sales_col else 0.0
      if total_sales == 0 and price_per_kg > 0 and weight > 0:
        total_sales = weight * price_per_kg

      total_sales_sum += total_sales

      phone_col = next((c for c in df.columns if 'هاتف' in c or 'عاتف' in c or 'phone' in c.lower()), None)
      phone = str(row.get(phone_col, "")).strip() if phone_col else ""
      if phone.endswith(".0"):
        phone = phone[:-2]
      phone = phone.replace("+", "").strip()
      if phone.startswith("964"):
        phone = phone[3:]
      formatted_phone = f"+964 {phone}" if phone and phone not in ["nan", "None"] else ""

      address_col = next((c for c in df.columns if 'عنوان' in c or 'address' in c.lower()), None)
      address = str(row.get(address_col, "")).strip() if address_col else ""
      if address in ["nan", "None"]:
        address = ""

      shipment_type = str(row.get(type_col_name, "")).strip() if type_col_name else ""
      if shipment_type in ["nan", "None"]:
        shipment_type = ""

      wb = openpyxl.load_workbook(active_template_file)
      ws = wb.active

      ws["B4"] = display_code
      ws["D4"] = today_date
      ws["B5"] = name
      ws["B6"] = address
      ws["D5"] = formatted_phone
      ws["B7"] = display_shipment
      ws["D6"] = packages
      ws["B8"] = shipment_type
      ws["D7"] = weight

      output = io.BytesIO()
      wb.save(output)
      output.seek(0)

      logo_img_tag = f'<img src="data:image/png;base64,{logo_base64}" style="max-height: 52px; max-width: 60px; margin-left: 10px; vertical-align: middle;">' if logo_base64 else ''

      single_receipt_html = f"""
            <div class="receipt-page" style="padding: 15px; font-family: 'Tahoma', Arial, sans-serif; direction: rtl; border: 2px solid #102a43; width: 100%; max-width: 148mm; margin: auto auto 20px auto; background: #ffffff; color: #102a43; box-sizing: border-box; page-break-after: always; break-after: page;">
                <table style="width: 100%; border-bottom: 2px solid #102a43; padding-bottom: 8px; margin-bottom: 12px; border-collapse: collapse;">
                    <tr>
                        <td style="text-align: right; vertical-align: middle;">
                            <div style="display: flex; align-items: center;">
                                {logo_img_tag}
                                <div>
                                    <h2 style="margin: 0; font-size: 15px; color: #102a43;">أطلس المحيط للتجارة العامة</h2>
                                    <p style="margin: 2px 0 0; font-size: 10px; color: #627d98;">OCEAN ATLAS GENERAL TRADING</p>
                                </div>
                            </div>
                        </td>
                        <td style="text-align: left; vertical-align: middle;">
                            <h3 style="margin: 0; font-size: 13px; color: #b45309;">وصل تسليم بضاعة</h3>
                            <p style="margin: 2px 0 0; font-size: 10px; color: #334e68;">Cargo Delivery Receipt</p>
                        </td>
                    </tr>
                </table>
                <table style="width: 100%; font-size: 11px; border-collapse: collapse; margin-bottom: 10px;">
                    <tr style="background-color: #f0f4f8;">
                        <td style="padding: 5px; border: 1px solid #bcccdc; width: 50%;"><strong>كود العميل:</strong> <span style="color: #b45309; font-weight: bold;">{display_code}</span></td>
                        <td style="padding: 5px; border: 1px solid #bcccdc; width: 50%;"><strong>رقم الشحنة:</strong> <span style="color: #b45309; font-weight: bold;">{display_shipment}</span></td>
                    </tr>
                    <tr>
                        <td style="padding: 5px; border: 1px solid #bcccdc;"><strong>اسم العميل:</strong> <span style="font-weight: bold;">{name}</span></td>
                        <td style="padding: 5px; border: 1px solid #bcccdc;"><strong>رقم الهاتف:</strong> <span style="direction: ltr; display: inline-block; font-weight: bold;">{formatted_phone}</span></td>
                    </tr>
                    <tr style="background-color: #f0f4f8;">
                        <td style="padding: 5px; border: 1px solid #bcccdc;"><strong>عنوان الاستلام:</strong> <span style="color: #486581; font-weight: bold;">{address}</span></td>
                        <td style="padding: 5px; border: 1px solid #bcccdc;"><strong>عدد الطرود:</strong> 📦 {packages} طرد</td>
                    </tr>
                    <tr>
                        <td style="padding: 5px; border: 1px solid #bcccdc;"><strong>تاريخ الإصدار:</strong> <span style="color: #b45309; font-weight: bold;">{today_date}</span></td>
                        <td style="padding: 5px; border: 1px solid #bcccdc;"><strong>الوزن الإجمالي:</strong> <span style="color: #102a43; font-weight: bold;">{weight} كغ</span></td>
                    </tr>
                    <tr style="background-color: #f0f4f8;">
                        <td style="padding: 5px; border: 1px solid #bcccdc;"><strong>نوع الشحنة:</strong> {shipment_type}</td>
                        <td style="padding: 5px; border: 1px solid #bcccdc;"><strong>حجم الشحنة (CBM):</strong> <span style="color: #b45309; font-weight: bold;">{cbm_value}</span></td>
                    </tr>
                    <tr>
                        <td style="padding: 5px; border: 1px solid #bcccdc;" colspan="2"><strong>سعر الكيلو:</strong> {price_per_kg:,.2f} $</td>
                    </tr>
                    <tr style="background-color: #fef3c7;">
                        <td style="padding: 5px; border: 1px solid #f59e0b;" colspan="2"><strong>إجمالي المبيعات:</strong> <span style="color: #b45309; font-weight: bold; font-size: 12px;">{total_sales:,.2f} $</span> &nbsp;&nbsp;&nbsp;&nbsp;|&nbsp;&nbsp;&nbsp;&nbsp; <strong>طريقة الدفع:</strong> [ &nbsp; ] نقداً &nbsp;&nbsp; [ &nbsp; ] أجل</td>
                    </tr>
                </table>
                <div style="background-color: #fffbeb; border: 1px solid #fde68a; padding: 6px; border-radius: 4px; margin-bottom: 10px;">
                    <p style="margin: 0; font-size: 10px; color: #92400e; line-height: 1.3;"><strong>إقرار الاستلام:</strong><br>أقر أنا الموقع أدناه، بأنني استلمت البضاعة والشحنة المذكورة أعلاه كاملة، وبحالة سليمة وممتازة، ومطابقة لكافة الأوزان والأوصاف المدونة.</p>
                </div>
                <table style="width: 100%; font-size: 11px; margin-top: 5px; margin-bottom: 10px;">
                    <tr>
                        <td style="width: 50%; padding: 2px;"><strong>اسم المستلم:</strong><br><br>............................................</td>
                        <td style="width: 50%; padding: 2px; text-align: left;"><strong>توقيع وختم المستلم:</strong><br><br>............................................</td>
                    </tr>
                </table>
                <div style="border-top: 1px dashed #bcccdc; margin-top: 10px; padding-top: 6px; text-align: center; font-size: 9.5px; color: #334e68;">
                    <span>📍 العنوان: بغداد - المنصور - تقاطع الواد</span><span style="margin: 0 10px;">|</span><span style="direction: ltr; display: inline-block;">📞 هاتف: 07858588899 / 07814518989</span>
                </div>
            </div>
            """

      all_receipts_html_for_print += single_receipt_html
      receipts_data_list.append({
          "index": index,
          "name": name,
          "code": display_code,
          "shipment": display_shipment,
          "total_sales": total_sales,
          "output": output,
          "file_name_id": file_name_id,
          "single_html": single_receipt_html,
      })

    # --- بطاقات الإحصائيات ---
    st.markdown("""
        <style>
        .metric-card-1 { background-color: #eff6ff; border: 1px solid #bfdbfe; padding: 15px; border-radius: 10px; text-align: center; }
        .metric-card-2 { background-color: #f0fdf4; border: 1px solid #bbf7d0; padding: 15px; border-radius: 10px; text-align: center; }
        .metric-card-3 { background-color: #f5f3ff; border: 1px solid #ddd6fe; padding: 15px; border-radius: 10px; text-align: center; }
        .metric-card-4 { background-color: #fffbeb; border: 1px solid #fde68a; padding: 15px; border-radius: 10px; text-align: center; }
        .metric-card-5 { background-color: #fdf2f8; border: 1px solid #fbcfe8; padding: 15px; border-radius: 10px; text-align: center; }
        </style>
    """, unsafe_allow_html=True)

    m1, m2, m3, m4, m5 = st.columns(5)
    with m1:
      st.markdown(f'<div class="metric-card-1"><p style="margin:0; color:#1e40af; font-weight:bold; font-size:14px;">👥 عدد العملاء</p><h3 style="margin:5px 0 0; color:#1e3a8a; font-size:20px;">{total_clients_count} عميل</h3></div>', unsafe_allow_html=True)
    with m2:
      st.markdown(f'<div class="metric-card-2"><p style="margin:0; color:#166534; font-weight:bold; font-size:14px;">📦 إجمالي الطرود</p><h3 style="margin:5px 0 0; color:#14532d; font-size:20px;">{total_packages_count} طرد</h3></div>', unsafe_allow_html=True)
    with m3:
      st.markdown(f'<div class="metric-card-3"><p style="margin:0; color:#5b21b6; font-weight:bold; font-size:14px;">📐 إجمالي الحجم</p><h3 style="margin:5px 0 0; color:#4c1d95; font-size:20px;">{total_cbm_sum:,.2f} CBM</h3></div>', unsafe_allow_html=True)
    with m4:
      st.markdown(f'<div class="metric-card-4"><p style="margin:0; color:#92400e; font-weight:bold; font-size:14px;">⚖️ الوزن الكلي</p><h3 style="margin:5px 0 0; color:#78350f; font-size:20px;">{total_weight_sum:,.2f} كغ</h3></div>', unsafe_allow_html=True)
    with m5:
      st.markdown(f'<div class="metric-card-5"><p style="margin:0; color:#9d174d; font-weight:bold; font-size:14px;">💰 المبلغ الإجمالي</p><h3 style="margin:5px 0 0; color:#831843; font-size:20px;">{total_sales_sum:,.2f} $</h3></div>', unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    st.subheader(f"📋 جدول تفاصيل الشحنة المعروضة: [{selected_shipment_filter}] - النوع: [{selected_type_filter}]")

    display_table_df = df.copy()
    display_table_df.insert(0, "التسلسل", range(1, len(display_table_df) + 1))
    table_html = display_table_df.to_html(classes="custom-table", index=False, escape=False)

    custom_table_styling = f"""
    <style>
        .custom-table-container {{
            max-height: 450px;
            overflow-x: auto;
            overflow-y: auto;
            border: 1px solid #bcccdc;
            border-radius: 8px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.05);
            margin-bottom: 20px;
        }}
        .custom-table {{
            width: 100%;
            border-collapse: collapse;
            font-family: 'Tahoma', Arial, sans-serif;
            font-size: 13px;
            direction: rtl;
            background-color: #ffffff;
            color: #102a43;
            white-space: nowrap;
        }}
        .custom-table th {{
            background-color: #102a43 !important;
            color: #ffffff !important;
            text-align: right;
            padding: 12px 15px;
            font-weight: bold;
            border-bottom: 2px solid #0b1e33;
            position: sticky;
            top: 0;
            z-index: 10;
        }}
        .custom-table td {{
            padding: 10px 15px;
            border-bottom: 1px solid #e2e8f0;
            text-align: right;
        }}
        .custom-table tr:nth-child(even) {{
            background-color: #f8fafc;
        }}
        .custom-table tr:hover {{
            background-color: #f1f5f9;
        }}
    </style>
    <div class="custom-table-container">
        {table_html}
    </div>
    """
    st.html(custom_table_styling)
    
    # --- زر تصدير الجدول إلى PDF فعال ومباشر (يعمل بضغطة زر واحدة دون إعادة تحميل) ---
    table_pdf_html_component = f"""
        <!DOCTYPE html>
        <html lang="ar" dir="rtl">
        <head>
            <meta charset="UTF-8">
            <style>
                @page {{ size: A4 landscape; margin: 10mm; }}
                body {{ font-family: 'Tahoma', Arial, sans-serif; direction: rtl; color: #102a43; margin: 0; padding: 0; background: transparent; }}
                .export-btn {{
                    background-color: #102a43;
                    color: white;
                    padding: 12px 24px;
                    border: none;
                    border-radius: 6px;
                    cursor: pointer;
                    font-weight: bold;
                    font-size: 14px;
                    display: inline-flex;
                    align-items: center;
                    gap: 8px;
                    box-shadow: 0 2px 4px rgba(0,0,0,0.1);
                    font-family: 'Tahoma', Arial, sans-serif;
                }}
                .export-btn:hover {{
                    background-color: #0b1e33;
                }}
            </style>
        </head>
        <body>
            <button class="export-btn" onclick="exportTablePDF()">📄 تصدير الجدول الحالي إلى PDF</button>
            <script>
                const tableContent = `{table_html.replace('`', '\\`').replace('$', '\\$')}`;
                const filterInfo = 'الشحنة: {selected_shipment_filter} | النوع: {selected_type_filter}';
                function exportTablePDF() {{
                    var w = window.open('', '', 'height=900,width=1200');
                    w.document.write(`
                        <!DOCTYPE html>
                        <html lang="ar" dir="rtl">
                        <head>
                            <meta charset="UTF-8">
                            <title>تقرير جدول الشحنات - أطلس</title>
                            <style>
                                @page {{ size: A4 landscape; margin: 10mm; }}
                                body {{ font-family: Tahoma, Arial, sans-serif; direction: rtl; color: #102a43; padding: 20px; }}
                                h2 {{ text-align: center; color: #102a43; margin-bottom: 20px; font-size: 18px; }}
                                table {{ width: 100%; border-collapse: collapse; font-size: 11px; margin-top: 10px; }}
                                th {{ background-color: #102a43 !important; color: #ffffff !important; text-align: right; padding: 10px; border: 1px solid #0b1e33; font-weight: bold; -webkit-print-color-adjust: exact; print-color-adjust: exact; }}
                                td {{ padding: 8px 10px; border: 1px solid #cbd5e1; text-align: right; }}
                                tr:nth-child(even) {{ background-color: #f8fafc; -webkit-print-color-adjust: exact; print-color-adjust: exact; }}
                            </style>
                        </head>
                        <body>
                            <h2>تقرير جدول تفاصيل الشحنة - شركة أطلس المحيط (${{filterInfo}})</h2>
                            ${{tableContent}}
                        </body>
                        </html>
                    `);
                    w.document.close();
                    w.focus();
                    setTimeout(() => {{ w.print(); w.close(); }}, 600);
                }}
            </script>
        </body>
        </html>
    """
    st.components.v1.html(table_pdf_html_component, height=55)
    st.markdown("---")

    master_payload = f"""
        <!DOCTYPE html>
        <html lang="ar" dir="rtl">
        <head><meta charset="UTF-8"></head>
        <body>
            <button style="background-color: #047857; color: white; padding: 14px 28px; border: none; border-radius: 8px; cursor: pointer; font-weight: bold; font-size: 16px; width: 100%; max-width: 500px; display: block; margin: 0 auto; box-shadow: 0 4px 6px rgba(0,0,0,0.1);" onclick="printAll()">🖨️ طباعة الوصولات المعروضة دفعة واحدة (مقاس A5)</button>
            <script>
                const content = `{all_receipts_html_for_print.replace('`', '\\`').replace('$', '\\$')}`;
                function printAll() {{
                    var w = window.open('', '', 'height=900,width=800');
                    w.document.write('<html><head><style>@page {{ size: A5; margin: 5mm; }} body {{ direction: rtl; font-family: Tahoma; }}</style></head><body>' + content + '</body></html>');
                    w.document.close();
                    w.focus();
                    setTimeout(() => {{ w.print(); w.close(); }}, 600);
                }}
            </script>
        </body>
        </html>
        """
    st.components.v1.html(master_payload, height=75)
    st.markdown("---")

    for item in receipts_data_list:
      with st.expander(f"📄 وصل العميل: {item['name']} | كود: {item['code'] or 'بدون'} | الشحنة: {item['shipment']} | الإجمالي: {item['total_sales']:,.2f} $"):
        st.download_button(
            label=f"📥 تنزيل إكسل الوصل",
            data=item["output"],
            file_name=f"Delivery_Receipt_{item['file_name_id']}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"dl_{item['index']}",
        )
        st.markdown("<br>", unsafe_allow_html=True)
        st.components.v1.html(
            f"""<div style="direction:rtl">{item['single_html']}</div><button style="background:#102a43;color:white;padding:12px 20px;border:none;border-radius:6px;cursor:pointer;font-weight:bold;margin-top:15px;" onclick="window.print()">🖨️ طباعة هذا الوصل</button>""",
            height=700,
            scrolling=True
        )
      st.markdown("---")

  except Exception as e:
    st.error(f"حدث خطأ أثناء معالجة الملفات: {e}")
else:
  st.info("الرجاء رفع ملف الشحنات وقالب الوصل وملف معلومات العملاء من الشريط الجانبي.")

